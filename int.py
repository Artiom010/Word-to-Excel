import os
import sys
import re
import sqlite3
import requests  
from tkinter import *
from tkinter import ttk, filedialog, messagebox
from docx import Document
from openpyxl.styles import Border, Side  
from openpyxl import Workbook  
from openpyxl.utils import get_column_letter
from tkinter.simpledialog import askstring

selected_currency = ""
selected_file_path = ""

def get_resource_path(relative_path):
    """
    Returnează calea absolută la fișierul .ico, compatibilă și cu .exe-ul generat de PyInstaller
    """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

ico_path = get_resource_path("icon.ico")


def is_table_header_or_separator(line: str) -> bool:
    """
    Verifică dacă linia este un antet de tabel sau separator, pe care vrem să-l excludem din Excel.
    """
    line_stripped = line.strip()
    if len(line_stripped) > 0 and all(ch in ['-', '.'] for ch in line_stripped):
        return True
    
    keywords = [
        "п/п",
        "Код",
        "Бар-код",
        "Наименование товара",
        "Кол-во",
        "Сумма отп.",
        "|товара|",
        "(USD)"
    ]
    for kw in keywords:
        if kw in line_stripped:
            return True

    return False

def extract_text_from_word(doc_path):
    """
    Extrage textul din documentul Word.
    Returnează o listă de paragrafe (linii) care nu sunt goale.
    """
    if not os.path.exists(doc_path):
        print(f"Fișierul {doc_path} nu a fost găsit.")
        return []
    try:
        doc = Document(doc_path)
        paragraphs = [para.text.strip() for para in doc.paragraphs if para.text.strip()]
        return paragraphs
    except Exception as e:
        print(f"Eroare la citirea fișierului Word: {e}")
        return []

def extract_invoice_blocks(lines):
    """
    Împarte lista de linii în blocuri pe baza liniei care începe cu "Заявка".
    Fiecare bloc reprezintă datele unei facturi (antet + date tabel + subsol).
    Înainte de a adăuga linia curentă la un bloc, verificăm dacă linia este un antet/separator de tabel
    pe care dorim să-l excludem (is_table_header_or_separator).
    """
    blocks = []
    current_block = []
    for line in lines:
        if is_table_header_or_separator(line):
            continue

        if line.startswith("Заявка"):
            if current_block:
                blocks.append(current_block)
                current_block = []
        current_block.append(line)

    if current_block:
        blocks.append(current_block)

    return blocks

def get_conversion_rate():
    """
    Preia rata de conversie USD -> RUB de pe https://www.cbpmr.net/kursval.php?lang=ru.
    Se caută în pagina HTML un pattern ipotetic pentru cursul de conversie.
    Dacă preluarea ratei eșuează, se returnează o valoare implicită de 16.1.
    """
    url = "https://www.cbpmr.net/kursval.php?lang=ru"
    try:
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            html = response.text
            match = re.search(r"USD.*?(\d+\.\d+)", html, re.IGNORECASE)
            if match:
                rate_str = match.group(1)
                return float(rate_str)
            else:
                print("Nu s-a putut găsi rata de conversie USD în pagina CBPMR.")
        else:
            print("Eroare la accesarea paginii CBPMR, status code:", response.status_code)
    except Exception as e:
        print("Eroare la obținerea ratei de conversie:", e)
    return 16.1

def process_invoice_block(block_lines):
    """
    Procesează un bloc de factură, extrăgând rândurile tabelului cu datele facturii.
    """
    pattern = re.compile(r'^\s*(\d+)\.\s+(\d+)\s+(\d+)\s+(.+?)\s+(\d+\.\d{3})\s+(\d+\.\d{3})\s+(\d+\.\d{3})\s*$')
    
    table_rows = []
    for line in block_lines:
        m = pattern.match(line)
        if m:
            table_rows.append(m.groups()[1:])
    
    conversion_rate = None
    if selected_currency == "₽":
        conversion_rate = get_conversion_rate()
    
    for i, row in enumerate(table_rows):
        row_list = list(row)
        if selected_currency == "$":
            row_list[4] = f"${row_list[4]}"
            row_list[5] = f"${row_list[5]}"
        elif selected_currency == "₽":
            try:
                order_price_usd = float(row_list[4])
                total_usd = float(row_list[5])
                order_price_rub = order_price_usd * conversion_rate
                total_rub = total_usd * conversion_rate
                row_list[4] = f"₽{order_price_rub:.4f}"
                row_list[5] = f"₽{total_rub:.4f}"
            except Exception as e:
                print("Eroare la conversia valutelor:", e)
        table_rows[i] = tuple(row_list)
    
    return {"table": table_rows}

def create_db_table():
    conn = sqlite3.connect('facturi.db')
    cursor = conn.cursor()

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS produse (
        ID INTEGER PRIMARY KEY AUTOINCREMENT,
        IDPRODUCER TEXT,
        BARCODE TEXT,
        ROPRODUCTNAME TEXT,
        QUANTITYINBOX TEXT,
        ORDERPRICE TEXT,
        TOTAL TEXT
    )
    ''')

    conn.commit()
    conn.close()


def insert_data_into_db(invoice_data):
    """
    Adaugă datele extrase într-o bază de date SQLite (facturi.db).
    """
    conn = sqlite3.connect('facturi.db')
    cursor = conn.cursor()

    for row in invoice_data["table"]:
        try:
            cursor.execute(''' 
                INSERT INTO produse (IDPRODUCER, BARCODE, ROPRODUCTNAME, QUANTITYINBOX, ORDERPRICE, TOTAL)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', row)
            conn.commit()
        except sqlite3.Error as e:
            print(f"Eroare la inserarea datelor: {e}")

    conn.close()

def load_imported_data_from_db():
    """
    Încarcă doar datele importate din baza de date SQLite și le afișează într-un tabel Tkinter.
    """
    conn = sqlite3.connect('facturi.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM produse ORDER BY IDPRODUCER DESC")  
    rows = cursor.fetchall()
    conn.close()

    return rows

def list_files_in_directory():
    """
    Afișează fișierele .docx din directorul C:\\Users\\Administrator\\Desktop în fereastra Tkinter.
    """
    directory = "C:\\Users\\Administrator\\Desktop"
    files = [f for f in os.listdir(directory) if f.endswith('.docx')]
    return files

def on_file_select(event):
    """
    Selectează fișierul și procesează-l (la dublu clic pe fișier).
    """
    global selected_file_path
    selected_file = file_listbox.get(file_listbox.curselection())
    selected_file_path = os.path.join("C:\\Users\\Administrator\\Desktop", selected_file)
    preview_and_import_file(selected_file_path)

def save_data_to_excel(invoice_blocks, file_path):
    """
    Salvează datele extrase din blocurile de facturi într-un fișier Excel (.xlsx),
    unde fiecare factură este plasată într-o foaie separată.
    Fișierul va avea același nume ca documentul Word și va fi salvat într-un subfolder 'xlsx'
    din directorul C:\\Users\\Administrator\\Desktop.
    """
    subfolder = os.path.join("C:\\Users\\Administrator\\Desktop\\dox", "xls")
    os.makedirs(subfolder, exist_ok=True)
    
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    excel_file_path = os.path.join(subfolder, base_name + ".xls")
    
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    headers = ["IDPRODUCER", "BARCODE", "ROPRODUCTNAME", "QUANTITYINBOX", "ORDERPRICE", "TOTAL"]
    
    for idx, block in enumerate(invoice_blocks, start=1):
        if isinstance(block, dict):
            invoice_number = block.get("invoice_number")
            block_lines = block.get("lines", [])
        else:
            block_lines = block
            invoice_number = None
            if block_lines and isinstance(block_lines[0], str) and block_lines[0].startswith("Заявка"):
                import re
                match = re.search(r"Заявка\s*№\s*(\d+)", block_lines[0])
                if match:
                    invoice_number = match.group(1)
        
        if invoice_number:
            sheet_name = f"Заявка № {invoice_number}"
        else:
            sheet_name = f"Factura {idx}"
        sheet_name = sheet_name[:31]
        
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        
        invoice_data = process_invoice_block(block_lines)
        for row in invoice_data["table"]:
            ws.append(row)
        
        thin_side = Side(border_style="thin", color="000000")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
       
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    cell_length = len(str(cell.value)) if cell.value is not None else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            adjusted_width = max_length + 2 
            ws.column_dimensions[column].width = adjusted_width

    try:
        wb.save(excel_file_path)
        messagebox.showinfo("Excel salvat", f"Fișierul Excel a fost salvat în: {excel_file_path}")
    except Exception as e:
        messagebox.showerror("Eroare Excel", f"Eroare la salvarea fișierului Excel: {e}")

def select_currency_and_import(invoice_blocks, file_path):
    """
    Deschide o fereastră pentru a selecta valuta și, după confirmare, procesează și inserează datele.
    """
    currency_window = Toplevel(root)
    currency_window.title("Selectează Valuta")

    Label(currency_window, text="Selectează valuta pentru prețuri:").pack(pady=10)
    currency_var = StringVar(value="$")  

    Radiobutton(currency_window, text="USD ($)", variable=currency_var, value="$").pack(pady=5)
    Radiobutton(currency_window, text="RUB (₽)", variable=currency_var, value="₽").pack(pady=5)

    def on_select():
        global selected_currency
        selected_currency = currency_var.get()
        currency_window.destroy()
        
        for block in invoice_blocks:
            invoice_data = process_invoice_block(block)
            insert_data_into_db(invoice_data)
        messagebox.showinfo("Succes", "Datele au fost importate cu succes în baza de date.")
        refresh_data()
        
        save_data_to_excel(invoice_blocks, file_path)

    Button(currency_window, text="Confirmă", command=on_select).pack(pady=10)

def preview_and_import_file(file_path):
    """
    Verifică fișierul Word și afișează o previzualizare înainte de importul datelor în SQLite.
    """
    text_data = extract_text_from_word(file_path)
    
    if text_data:
        invoice_blocks = extract_invoice_blocks(text_data)
        
        if invoice_blocks:
            preview_window = Toplevel(root)
            preview_window.title("Preview Facturi")

            preview_treeview = ttk.Treeview(preview_window, columns=("ID", "Barcode", "Produs", "Cantitate", "Preț", "Total"), show="headings")
            preview_treeview.heading("ID", text="ID")
            preview_treeview.heading("Barcode", text="Barcode")
            preview_treeview.heading("Produs", text="Produs")
            preview_treeview.heading("Cantitate", text="Cantitate")
            preview_treeview.heading("Preț", text="Preț")
            preview_treeview.heading("Total", text="Total")
            preview_treeview.pack(padx=10, pady=10)

            for i, block in enumerate(invoice_blocks, start=1):
                invoice_data = process_invoice_block(block)
                for row in invoice_data["table"]:
                    preview_treeview.insert("", "end", values=row)

            def on_confirm():
                preview_window.destroy()  
                select_currency_and_import(invoice_blocks, file_path) 

            def on_cancel():
                preview_window.destroy()

            Button(preview_window, text="Confirmă Importul", command=on_confirm).pack(pady=10, side=LEFT, padx=20)
            Button(preview_window, text="Anulează", command=on_cancel).pack(pady=10, side=RIGHT, padx=20)
        else:
            messagebox.showerror("Eroare", "Nu s-au găsit blocuri de facturi în document.")
    else:
        messagebox.showerror("Eroare", "Nu au fost găsite date de salvat.")

def refresh_data():
    """
    Încarcă și afișează doar datele importate din baza de date SQLite în tabelul Tkinter.
    """
    rows = load_imported_data_from_db()  
    for row in treeview.get_children():
        treeview.delete(row) 
    for row in rows:
        treeview.insert("", "end", values=row)

def clear_table():
    """
    Șterge toate datele importate din tabelul Tkinter și din baza de date,
    dar numai dacă se introduce parola corectă.
    """
    password = askstring("Verificare Parolă", "Introdu parola:", show="*")
    if password != "pased":
        messagebox.showerror("Eroare", "Parola incorectă!")
        return

    if messagebox.askyesno("Confirmare", "Dorești să ștergi toate datele importate?"):
        conn = sqlite3.connect('facturi.db')
        cursor = conn.cursor()
        cursor.execute("DELETE FROM produse")  
        conn.commit()
        conn.close()
        
        for row in treeview.get_children():
            treeview.delete(row)

        messagebox.showinfo("Succes", "Toate datele au fost șterse din tabel.")

root = Tk()
root.title("Import Facturi")
root.iconbitmap(ico_path)



Label(root, text="Selectează fișierul Word").pack(pady=10)

file_listbox = Listbox(root, width=50, height=10)
file_listbox.pack(pady=10)

files = list_files_in_directory()
for file in files:
    file_listbox.insert(END, file)

file_listbox.bind("<Double-1>", on_file_select)

treeview = ttk.Treeview(root, columns=("ID", "Barcode", "Nume Produs", "Cantitate in Cutie", "Pret Unit", "Total"), show="headings")
treeview.heading("ID", text="ID")
treeview.heading("Barcode", text="Barcode")
treeview.heading("Nume Produs", text="Nume Produs")
treeview.heading("Cantitate in Cutie", text="Cantitate in Cutie")
treeview.heading("Pret Unit", text="Pret Unit")
treeview.heading("Total", text="Total")
treeview.pack(padx=10, pady=10)

clear_button = Button(root, text="X", command=clear_table, fg="red", font=("Arial", 16), relief=FLAT, width=3, height=1)
clear_button.place(x=20, y=20)  

create_db_table()
refresh_data()
root.mainloop()
